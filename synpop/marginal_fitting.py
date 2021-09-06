"""
This module contains the framework to modify agents and match population wide control totals.
"""
from typing import List
import logging
import warnings
import math
from pathlib import Path

import pandas as pd
from openpyxl.styles.fonts import Font
from openpyxl import load_workbook

# fix the random seed for the marginal fitting process
import numpy
from synpop.config import RandomFittingConfig
from synpop import visualisations
from synpop import zone_maps
numpy.random.seed(RandomFittingConfig().random_seed)

# Logger settings set in __init__.py
logger = logging.getLogger(__name__)


# Building marginal tables ##
def compute_comparison_summary(df_persons, df_persons_ist, year, year_ist, groupby):
    """
    Computes a comparison table of the two populations aggregated by the feature or features given.
    """
    counts = df_persons.groupby(groupby).count().iloc[:, 0].rename('counts_{}'.format(str(year)))
    counts_ist = df_persons_ist.groupby(groupby).count().iloc[:, 0].rename('counts_{}'.format(str(year_ist)))

    growth = (counts / counts_ist).round(2).rename('growth')

    prop = (counts / df_persons.shape[0]).round(2).rename('prop_{}'.format(str(year)))
    prop_ist = (counts_ist / df_persons_ist.shape[0]).round(2).rename('prop_{}'.format(str(year_ist)))

    return pd.concat([counts_ist, counts, growth, prop_ist, prop], axis=1).fillna(0).round(2)


def create_summary_table(df_persons, df_persons_ist, year, year_ist, groupby):
    # same as previous method but with a nicer formatting for display
    summary_table = compute_comparison_summary(df_persons, df_persons_ist, year, year_ist, groupby)

    summary_table['growth'] = summary_table['growth'] - 1
    summary_table.columns = [f'counts {year_ist}', f'counts {year}',
                             'absolute growth (%)', f'proportion {year_ist}', f'proportion {year}']
    summary_table[['absolute growth (%)', f'proportion {year_ist}', f'proportion {year}']] = (
            100 * summary_table[['absolute growth (%)', f'proportion {year_ist}', f'proportion {year}']])
    summary_table.index.name = groupby
    return summary_table

def compute_ist_vs_scenario_marginal_counts(df, df_ist, year, year_ist, feature, control_level_list):
    """
    The scenario and ist population are aggregated by the feature and the control variables.
    --> The rows are counted in each aggregation group
    """
    # Copies to make sure original data is not touched
    df = df.copy(deep=True)
    df_ist = df_ist.copy(deep=True)

    group_by = control_level_list + [feature, ]

    # Changing boolean to string to avoid boolean as index and columns
    if df[feature].dtype == bool:
        df[feature] = df[feature].replace({False: 'false', True: 'true'})
        df_ist[feature] = df_ist[feature].replace({False: 'false', True: 'true'})

    # If feature is not categorical, make it. This will keep the empty categories when grouping by
    if df[feature].dtype.name != 'category':
        df[feature] = df[feature].astype('category')
        df_ist[feature] = df_ist[feature].astype('category')

    counts_ist = (df_ist.groupby(group_by).count().iloc[:, 0]
                  .fillna(0).astype(int)
                  .rename('counts_{}'.format(year_ist))
                  )

    counts = (df.groupby(group_by).count().iloc[:, 0]
              .fillna(0).astype(int)
              .rename('counts_{}'.format(year))
              )

    marginals = pd.concat((counts_ist, counts), axis=1, sort=True).fillna(0).astype(int)
    return marginals


def compute_ist_vs_scenario_marginal_summary_table(df, df_ist, year, year_ist, feature, control_level_list):
    """
    The scenario and ist population are aggregated by the feature and the control variables.
    --> The rows are counted in each aggregation group
    --> The ratio of counts per total are computed for the IST year
    --> The ratio are applied to the scenario to compute the expected counts
    --> The deltas are the differences between expected and actual counts
    """
    # Copies to make sure original data is not touched
    df = df.copy(deep=True)
    df_ist = df_ist.copy(deep=True)

    # Computing counts
    counts_df = compute_ist_vs_scenario_marginal_counts(df, df_ist, year, year_ist, feature, control_level_list)
    counts_ist = counts_df['counts_{}'.format(year_ist)]
    counts = counts_df['counts_{}'.format(year)]

    group_by = control_level_list + [feature, ]

    # Computing ratio in the IST year
    counts_per_pop_segment_ist = (counts_ist
                                  .reset_index()
                                  .groupby(control_level_list)['counts_{}'.format(year_ist)].sum()
                                  )

    ratios_ist = (counts_ist / counts_per_pop_segment_ist).rename('ratios_{}'.format(year_ist))

    # Computing the expected counts in the scenario year
    counts_per_pop_segment = df.groupby(control_level_list)['person_id'].count()
    expected_counts = ((ratios_ist * counts_per_pop_segment)
                       .fillna(0)
                       .astype(int)
                       .rename('expected_counts_{}'.format(year))
                       )
    # When there are zero counts in a population segment in IST, there can be no valid predictions.
    # In that case, the expected counts are set to the actual counts to avoid correcting based on this artifact.
    segments_with_no_ist_data = counts_per_pop_segment_ist[counts_per_pop_segment_ist == 0].index
    if len(segments_with_no_ist_data) > 0:
        expected_counts.loc[segments_with_no_ist_data] = counts.loc[segments_with_no_ist_data]
        logger.info('The following segments have no data in ist and will stay untouched: {}'
                       .format(segments_with_no_ist_data)
                       )

    deltas = (counts - expected_counts).rename('deltas_{}'.format(year))

    marginals = (pd.concat((counts_ist, ratios_ist, counts, expected_counts, deltas), axis=1, sort=True)
                 .fillna(0)
                 .astype({'counts_{}'.format(year): int, 'counts_{}'.format(year_ist): int,
                          'expected_counts_{}'.format(year): int, 'deltas_{}'.format(year): int
                          })
                 )

    return marginals


def compute_bfs_vs_scenario_marginal_summary_table(df, bfs_ratio_prediction, year, feature, control_level_list):
    """
    The scenario and ist population are aggregated by the feature and the control variables.
    --> The rows are counted in each aggregation group
    --> The ratio of counts per total are computed for the IST year
    --> The ratio are applied to the scenario to compute the expected counts
    --> The deltas are the differences between expected and actual counts
    """
    # Copies to make sure original data is not touched
    df = df.copy(deep=True)

    if type(control_level_list) is not list:
        control_level_list = [control_level_list]

    group_by = control_level_list + [feature, ]

    # Change types to category to get empty  options in the groupby as well
    original_types = df[group_by].dtypes.to_dict()  # keep this to set original types back after processing
    df[group_by] = df[group_by].astype('category')
    counts = df.groupby(group_by).count().iloc[:, 0].rename('counts_{}'.format(year)).fillna(0).astype(int)

    bfs_ratio_prediction = (bfs_ratio_prediction
                            .reindex(counts.index)
                            .fillna(0)
                            .rename('expected_ratios_{}'.format(year))
                            )

    counts_per_category = counts.reset_index().groupby(control_level_list).sum()['counts_{}'.format(year)]
    expected_counts = ((bfs_ratio_prediction * counts_per_category)
                       .fillna(0)
                       .astype(int)
                       .rename('expected_counts_{}'.format(year))
                       )

    deltas = (counts - expected_counts).rename('deltas_{}'.format(year))

    marginals = (pd.concat((bfs_ratio_prediction, counts, expected_counts, deltas), axis=1, sort=True).fillna(0))
    marginals = marginals.astype({'counts_{}'.format(year): int,
                                  'expected_counts_{}'.format(year): int,
                                  'deltas_{}'.format(year): int})
    # Setting index types to original
    marginals = marginals.reset_index().astype(original_types).set_index(group_by)
    return marginals


def compute_cross_table(df, main_feature, secondary_features):
    # Copies to make sure original data is not touched
    df = df.copy(deep=True)

    if type(secondary_features) is not list:
        secondary_features = [secondary_features]

    counts = (df.groupby(secondary_features + [main_feature]).count().iloc[:, 0]
              .fillna(0).astype(int)
              .rename('counts')
              .reset_index()
              )

    cross_table = counts.pivot_table(index=secondary_features, columns=main_feature, values='counts', fill_value=0)

    # In case the index is boolean, it is converted to string
    str_indices = []
    for i in range(cross_table.index.nlevels):
        str_indices.append(cross_table.index.get_level_values(i).astype(str))

    cross_table.index = str_indices

    return cross_table


# Fixing marginals ##
def fix_categorical_feature(persons: pd.DataFrame, feature, pop_segment_variables, control_totals: pd.DataFrame,
                            person_proba: pd.DataFrame = None):
    """
    The idea is not to pick the persons to change compleatly at random but based on a probabilistic model that gives
    and estimation of the likelihood that person is in the wrong category.
    This functionality is implemented for categorical features.
    persons.

    :param persons:
    :param feature:
    :param pop_segment_variables:
    :param control_totals:
    :param person_proba:
    """
    # Copies to make sure original data is not touched
    persons = persons.copy(deep=True)

    # If only a string is given, cast pop_segment_variables into a list
    if type(pop_segment_variables) is not list:
        pop_segment_variables = [pop_segment_variables]

    # If no probabilities are given, a uniform distribution is used.
    if person_proba is None:
        logger.info('Since no probability model has been given, a uniform distribution will be used.')
        person_proba = build_person_proba_with_uniform_distribution(persons, persons[feature].unique())

    # Change boolean to text to avoid weird issues. If it is already text, nothing will happen.
    persons[feature] = persons[feature].replace({False: 'false', True: 'true'})
    control_totals = control_totals.rename(columns={False: 'false', True: 'true'})
    person_proba = person_proba.rename(columns={False: 'false', True: 'true'})

    # Some sanity checks
    assert person_proba.shape[0] == persons.shape[0], 'Must have a probabilities for all persons'
    # Count persons in all population segments
    raw_pop_counts = (persons
                      .groupby(pop_segment_variables + [feature]).count()
                      .iloc[:, 0]
                      .fillna(0).astype(int)
                      .rename('counts')
                      .reset_index()
                      .pivot_table(index=pop_segment_variables, columns=feature, values='counts', fill_value=0)
                      )

    # raw + delta = control total
    delta_counts = (control_totals - raw_pop_counts).fillna(0).astype(int)

    feature_fixed = persons.set_index('person_id')[feature].copy(deep=True)

    logger.info('Fixing "{}" by population segments based on: {}'.format(feature, pop_segment_variables))
    tot_pop_segments = persons[pop_segment_variables].drop_duplicates().shape[0]

    # Group by the population is important to efficiently segment the population with n variables
    i = 0
    for segment_name, persons_segment in persons.groupby(pop_segment_variables):
        i += 1

        deltas = delta_counts.loc[segment_name]
        logger.debug('Sampling changes for segment = {} ({})'.format(segment_name, deltas.to_dict()))

        sampled_changes = _pick_the_persons_to_change(persons_segment.set_index('person_id')[feature],
                                                      person_proba,
                                                      deltas
                                                      )

        for cat, ids in sampled_changes.items():
            feature_fixed.loc[ids] = cat

        if (i % 20) == 0:
            logger.info('{} / {} population segments fixed...'.format(i, tot_pop_segments))

    logger.info('All {} population segments fixed!'.format(i))

    # fix boolean attribute
    if feature_fixed.drop_duplicates().isin(['true', 'false']).all():
        feature_fixed = feature_fixed.replace({'false': False, 'true': True})

    return feature_fixed


def _pick_the_persons_to_change(persons_raw, person_probabilities, deltas):
    """
    # Observation: this method takes close to 1 second and is the bottleneck in terms of computation time

    To optimise computational time, the agents are not sampled and modified one by one.
    This function sample a pool of agents from all categories with too many people.
    Each person in this pool of people is then assigned one of the categories with too few people.
    The modifications are done in one go afterwards.
    """
    # raw + delta = control total
    cats_with_too_many_people = deltas[deltas < 0].index.values
    logger.debug(f'The categories with too many people are: {cats_with_too_many_people}')

    # Check if there are any people to change
    total_changes = abs(deltas.loc[cats_with_too_many_people].sum())
    if total_changes == 0:
        logger.debug(f'The total number of peoples to change is zero. Returning ... ')
        return dict()  # no changes

    cats_with_too_few_people = deltas[deltas > 0].index.values
    logger.debug(f'The categories with too few people are: {cats_with_too_few_people}')

    # The correct number of people are pre-sampled from each category with too many people.
    # This forms the pool of persons that can be modified.
    person_subpools = []
    sampled_changes = dict()
    for origin_cat in cats_with_too_many_people:
        with warnings.catch_warnings():
            # This will hide one confusing FutureWarning message that should not be sent to the user
            # ref: https://stackoverflow.com/a/46721064
            warnings.simplefilter(action='ignore', category=FutureWarning)
            persons_in_origin_cat = persons_raw[persons_raw == origin_cat]

        logger.debug(f'Cat. "{origin_cat}" contains {len(persons_in_origin_cat)} persons')
        logger.debug(f'{-deltas[origin_cat]} persons from "{origin_cat}" will be sampled out')

        # The probability of being chose for change from A is the probability of not being A (1 - P(A))
        weights = 1 - person_probabilities.loc[persons_in_origin_cat.index, origin_cat]
        persons_to_change = persons_in_origin_cat.sample(n=-deltas[origin_cat],
                                                         replace=False,
                                                         weights=weights
                                                         )

        person_subpools.append(persons_to_change)
        logger.debug(f'{len(persons_to_change)} persons with original cat = "{origin_cat}" have been sampled')

    person_pool = pd.concat(person_subpools, axis=0)
    logger.debug(f'Person-pool has been filled up with {len(person_pool)} persons')

    for dest_cat in cats_with_too_few_people:
        # Pick randomly the persons from the pool weighted with their probability of being in dest
        persons_to_sample = min(deltas[dest_cat], person_pool.shape[0])  # to avoid rounding induced bugs
        diff = abs(persons_to_sample - deltas[dest_cat])
        assert diff < 10, 'This difference should always be very small! (diff = {})'.format(diff)

        logger.debug('Picking {} persons for "{}"'.format(persons_to_sample, dest_cat))

        # There is bug with sample happens when sampling a big proportion of values with weights and without replacement
        # Solution: if sampling more than 50%, the persons to be left out are sampled instead
        if persons_to_sample <= (person_pool.shape[0] * 0.5):
            logger.debug(f'Sampling directly {persons_to_sample} to change')
            sampled_persons = (person_pool
                               .sample(n=persons_to_sample,
                                       replace=False,
                                       weights=person_probabilities.loc[person_pool.index, dest_cat]
                                       )
                               .index
                               )
        else:
            persons_not_to_sample = person_pool.shape[0] - persons_to_sample
            logger.debug(f'Sampling indirectly {persons_not_to_sample} people not to change (to avoid a numpy bug...)')
            persons_not_sampled = (person_pool
                                   .sample(n=persons_not_to_sample,
                                           replace=False,
                                           weights=1 - person_probabilities.loc[person_pool.index, dest_cat]
                                           )
                                   .index
                                   )

            sampled_persons = person_pool.index[~person_pool.index.isin(persons_not_sampled)]

        sampled_changes[dest_cat] = sampled_persons
        logger.debug(f'{len(sampled_persons)} people sampled to be converted to "{dest_cat}"')
        logger.debug(f'Total number of people sampled for change: {len(sampled_changes[dest_cat])}')

        # Remove the sampled person from the person pool (redefining is much faster than dropping rows)
        person_pool = person_pool[~person_pool.index.isin(sampled_persons)]
        logger.debug(f'Person-pool now contains {len(person_pool)} persons')
    final_report = {f'to "{v}"': len(k) for v, k in sampled_changes.items()}
    logger.debug(f'All sampling done: {final_report}')
    return sampled_changes


def build_person_proba_with_uniform_distribution(persons, categories):
    person_proba = pd.DataFrame(1, index=persons['person_id'], columns=categories).fillna(1)
    person_proba = person_proba / len(categories)
    person_proba.index.name = 'person_id'
    assert (person_proba.sum(axis=1) - 1.0 < 1e-8).all()
    return person_proba


def build_person_proba_with_from_cross_table(persons, cross_table, merge_on):
    # Copies to make sure original data is not touched
    persons = persons.copy(deep=True)

    # If only a string is given, cast it into a list
    if type(merge_on) is not list:
        merge_on = [merge_on]

    # In to facilitate joining, convert all to string
    persons[merge_on] = persons[merge_on].astype(str)

    # All probabilities can be zero for one line if there was no counts for that category in the cross table.
    # This should very rarely happen, but when it does, a uniform probability is given to all possibilities
    uniform_proba = 1 / cross_table.shape[1]

    probabilities = cross_table.div(cross_table.sum(axis=1), axis=0).fillna(uniform_proba)

    person_proba = (pd.merge(persons[['person_id'] + merge_on], probabilities, left_on=merge_on, right_index=True)
                    .drop(merge_on, axis=1)
                    .set_index('person_id')
                    )

    assert person_proba.sum(axis=1).apply(lambda v: math.isclose(v, 1)).all()

    # There are some issues with sampling when some probabilities are exactly zero. This makes it work:
    person_proba = person_proba.mask(person_proba == 0, 1e-10)
    person_proba = person_proba.mask(person_proba == 1, 1 - 1e-10)

    return person_proba


def write_fitting_goal(excel_path, expected_counts, target_variable, fitting_segments, proba_weights='none/uniform',
                       overwrite=False):
    mode = 'a' if Path(excel_path).exists() and not overwrite else 'w'
    if isinstance(fitting_segments, list):
        fitting_segments = ','.join(fitting_segments)

    if isinstance(proba_weights, list):
        proba_weights = ','.join(proba_weights)

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode) as writer:
        # write control totals
        expected_counts.to_excel(writer, sheet_name=target_variable)

        # insert new rows. If there are merged cells (MultiIndex), shift them first
        sheet = writer.sheets[target_variable]
        merged_cells_range = sheet.merged_cells.ranges
        for merged_cell in merged_cells_range:
            merged_cell.shift(0, 4)
        sheet.insert_rows(idx=0, amount=4)

        # fill config and format
        sheet['A1'] = 'target_variable'
        sheet['A2'] = 'fitting_segments'
        sheet['A3'] = 'probability_weights'
        sheet['B1'] = target_variable
        sheet['B2'] = fitting_segments
        sheet['B3'] = proba_weights
        sheet['A1'].font = Font(bold=True)
        sheet['A2'].font = Font(bold=True)
        sheet['A3'].font = Font(bold=True)
        sheet.column_dimensions['A'].width = 20


def parse_fitting_goals(excel_path):
    wb = load_workbook(excel_path)
    configs = {}
    for target_variable in wb.sheetnames:
        # parse config
        config = pd.read_excel(excel_path, sheet_name=target_variable, engine='openpyxl', usecols=[0, 1], header=None,
                               nrows=3, names=['config_key', 'config_value'], index_col=0)['config_value']
        config.loc['fitting_segments'] = config.loc['fitting_segments'].split(',')
        config.loc['probability_weights'] = config.loc['probability_weights'].split(',')
        if config.loc['probability_weights'] == ['none/uniform']:
            config.loc['probability_weights'] = None

        # parse target table (index_col parameter can handle MultiIndex)
        expected_counts = pd.read_excel(excel_path, sheet_name=target_variable, engine='openpyxl', skiprows=4,
                                        index_col=list(range(len(config.loc['fitting_segments']))))

        configs[target_variable] = (config, expected_counts)

    return configs


def fit_target_variable(persons, persons_ist, year, year_ist, fitting_config, expected_counts, plots_output_path=None):
    # Wrapper function which does all the fitting-related work and returns the full transformed persons dataframe.
    persons_fixed = persons.copy(deep=True)
    target_variable = fitting_config['target_variable']

    # prepare person probabilities if required
    if fitting_config['probability_weights'] is not None:
        if persons_ist is None:
            raise ValueError("For weighted fits (probability_weights != None), "
                             "'persons_ist' must be provided for the cross-table")
        cross_table = compute_cross_table(persons_ist,
                                          main_feature=target_variable,
                                          secondary_features=fitting_config['probability_weights']
                                          )
        person_proba = build_person_proba_with_from_cross_table(
            persons_fixed, cross_table,
            merge_on=fitting_config['probability_weights']
        )
    else:
        # uniform probability is used
        person_proba = None

    # fix variable with regular marginal fitting
    fixed_variable = fix_categorical_feature(persons=persons_fixed,
                                             feature=target_variable,
                                             pop_segment_variables=fitting_config['fitting_segments'],
                                             control_totals=expected_counts,
                                             person_proba=person_proba
                                             )

    # update persons dataframe and return
    persons_fixed = persons_fixed.drop(target_variable, axis=1)
    persons_fixed = pd.merge(persons_fixed, fixed_variable, left_on='person_id', right_index=True)

    if plots_output_path is not None:
        # TODO: convert this function into a class for clarity and flexibility
        # outputs location
        path = Path(plots_output_path) / target_variable
        path.mkdir(exist_ok=True, parents=True)

        # store summary tables as CSV
        create_summary_table(persons, persons_ist, year, year_ist, target_variable
                             ).to_csv(path / "summary_table_raw.csv")
        create_summary_table(persons_fixed, persons_ist, year, year_ist, target_variable
                             ).to_csv(path / "summary_table_fixed.csv")
        is_binary_feature = persons[target_variable].drop_duplicates().isin([True, False]).all()

        for _persons, _year in zip([persons, persons_fixed, persons_ist], [f'{year}-Raw', year, year_ist]):
            is_prognose_year = int(''.join(c for c in str(_year) if c.isnumeric())) > 2020
            if (target_variable in ('current_edu', 'current_job_rank', 'is_employed') and
                    fitting_config['probability_weights'] is not None):
                # generate cross-tables plots
                _ = visualisations.plot_ct1(persons, title=f'SynPop{_year}: Education vs Employment')
                visualisations.save_figure(True, f'SynPop{_year}_CT1.png', path)
                _ = visualisations.plot_ct2(persons, title=f'SynPop{_year}: Education vs Job-Rank')
                visualisations.save_figure(True, f'SynPop{_year}_CT2.png', path)
                _ = visualisations.plot_ct3(persons, title=f'SynPop{_year}: Job-Rank vs Employment')
                visualisations.save_figure(True, f'SynPop{_year}_CT3.png', path)

            if 'age' in fitting_config['fitting_segments']:
                # generate pyramid plots
                counts_per_cat = _persons.groupby(
                    ['age', target_variable])['person_id'].count().unstack(level=1)
                if is_binary_feature:
                    counts_per_cat = counts_per_cat.rename(columns={True: 'true', False: 'false'}, errors='ignore')
                    counts_per_cat = counts_per_cat[['true', 'false']]

                ax = None
                if is_binary_feature and is_prognose_year:
                    # Expected values based on BFS-predictions
                    ax = (expected_counts.groupby('age')[True].sum().rename('Expected')
                          .replace(0, numpy.nan).plot(style=':', marker='x', color='k', rot=0))
                    _ = ax.legend(loc='upper right')

                _ = visualisations.plot_multi_class_feature_per_age(
                    counts_per_cat, colour_dict=visualisations.COLOUR_DICTS[target_variable], ymax=150_000, ax=ax,
                    y_grid=is_binary_feature, title=f'{target_variable} by Age - SynPop{_year}')
                visualisations.save_figure(True, f'SynPop{_year}_{target_variable}_by_age.png', path)

            if 'KT_full' in fitting_config['fitting_segments'] and is_binary_feature and is_prognose_year:
                # generate map plots
                mapper = zone_maps.SwissZoneMap()
                by_canton = pd.concat([expected_counts.groupby('KT_full')[True].sum(),
                                       _persons.groupby('KT_full')[target_variable].sum()],
                                      keys=['Expected', 'SynPop'], axis=1, sort=False)
                by_canton['delta_abs'] = (by_canton['SynPop'] - by_canton['Expected'])
                by_canton['delta_pc'] = (((by_canton['SynPop'] / by_canton['Expected']) - 1) * 100).round(1)

                # Absolute differences
                title = f'SynPop{_year} vs. Expected: Absolute Diff of {target_variable}'
                round_up = lambda x: int(math.ceil(x / 1000.0)) * 1000
                scale_bound = round_up(by_canton['delta_abs'].abs().max())
                _ = mapper.draw_cantons(by_canton, 'delta_abs', vmin=-scale_bound, vmax=scale_bound, title=title)
                visualisations.save_figure(True, f'SynPop{_year}_{target_variable}_by_Canton_absolute.png', path)

                # Relative differences
                title = f'SynPop{_year} vs. Expected: Relative (%) Diff of {target_variable}'
                scale_bound = round(by_canton['delta_pc'].abs().max())
                _ = mapper.draw_cantons(by_canton, 'delta_abs', vmin=-scale_bound, vmax=scale_bound, title=title)
                visualisations.save_figure(True, f'SynPop{_year}_{target_variable}_by_Canton_relative.png', path)

    return persons_fixed
